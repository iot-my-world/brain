import os
import random
import re
from datetime import datetime
from datetime import timedelta
from math import pi, sin, cos, atan2, sqrt

from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

earthRadiusInKm = 6378.137

minimumRandomStartDate = datetime(
    year=2019,
    month=2,
    day=20,
)

maximumRandomStartDate = datetime(
    year=2019,
    month=12,
    day=25,
)


def difference_between_readings(r1, r2):
    lat1 = float(r1[0])
    lon1 = float(r1[1])
    lat2 = float(r2[0])
    lon2 = float(r2[1])
    dLat = lat2 * pi / 180 - lat1 * pi / 180
    dLon = lon2 * pi / 180 - lon1 * pi / 180
    a = sin(dLat / 2) * sin(dLat / 2) + cos(lat1 * pi / 180) * cos(lat2 * pi / 180) * sin(dLon / 2) * sin(dLon / 2)
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    d = earthRadiusInKm * c
    return d * 1000


def first_empty_column(worksheet):
    return get_column_letter(len(list(worksheet.columns)) + 1)


def next_column_letter(column_letter):
    return get_column_letter(column_index_from_string(column_letter) + 1)


def random_start_date(start=minimumRandomStartDate, end=maximumRandomStartDate):
    """Generate a random datetime between `start` and `end`"""
    return start + timedelta(
        # Get a random amount of seconds between `start` and `end`
        seconds=random.randint(0, int((end - start).total_seconds())),
    )


def get_journey_name(filepath):
    # ./raw/dbnCpt.rdat
    return filepath.split('/')[2].split('.')[0]


stringPattern = r'"(.*?)"'

if __name__ == '__main__':
    # get reading file paths
    os.chdir('/Users/bernardbussy/go/src/github.com/iot-my-world/brain/testing/tracker/reading/data')
    readingFilePaths = ['./raw/' + readingFile for readingFile in os.listdir('./raw') if readingFile.endswith('.rdat')]
    # open and set up workbook to save readings in
    outputWorkbook = Workbook()
    journeyName = readingFilePaths[0].split('.')[0]
    outputWorkbook.worksheets[0].title = get_journey_name(readingFilePaths[0])
    outputWorkbook.active = 0
    # for every readings data file
    for filePathIdx, filePath in enumerate(readingFilePaths):
        # add new sheet for each journey
        if filePathIdx > 0:
            journeyName = get_journey_name(filePath)
            outputWorkbook.create_sheet(journeyName)
            outputWorkbook.active = filePathIdx

        startDateTime = random_start_date()
        journeyTime = 0
        readings = []

        # open the file and get the readings and journey time
        with open(filePath) as readingFile:
            for line in readingFile:
                if 'TIME:' in line:
                    journeyTime = datetime.strptime(re.findall(stringPattern, line)[0], '%H:%M')

                if ('trkpt' in line) and ('lat' in line) and ('lon' in line):
                    readings.append(re.findall(stringPattern, line))

        # write out the readings with time stamps
        row = 1
        latColumn = first_empty_column(outputWorkbook.active)
        lonColumn = next_column_letter(latColumn)
        timeStampColumn = next_column_letter(lonColumn)
        outputWorkbook.active[latColumn + str(row)].value = "Lat"
        outputWorkbook.active[lonColumn + str(row)].value = "Lon"
        outputWorkbook.active[timeStampColumn + str(row)].value = "stamp"
        timestampDelta = timedelta(
            hours=journeyTime.hour / len(readings),
            minutes=journeyTime.minute / len(readings),
        )
        lastTimestamp = startDateTime
        lastReading = readings[0]
        row += 1
        for readingIdx, reading in enumerate(readings):
            if readingIdx != 0:
                if difference_between_readings(lastReading, reading) < 100:
                    continue
                else:
                    row += 1
                    lastReading = reading

            lastTimestamp += timestampDelta
            print(journeyName, reading)
            outputWorkbook.active[latColumn + str(row)].value = reading[0]
            outputWorkbook.active[lonColumn + str(row)].value = reading[1]
            outputWorkbook.active[timeStampColumn + str(row)].value = lastTimestamp.strftime("%s")

    outputWorkbook.save('data.xlsx')
    print('done')
