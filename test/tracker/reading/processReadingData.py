import os
import random
import re
from datetime import datetime
from datetime import timedelta

from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string


def first_empty_column(worksheet):
    return get_column_letter(len(list(worksheet.columns)) + 1)


def next_column_letter(column_letter):
    return get_column_letter(column_index_from_string(column_letter) + 1)


def random_date(start, end):
    """Generate a random datetime between `start` and `end`"""
    return start + timedelta(
        # Get a random amount of seconds between `start` and `end`
        seconds=random.randint(0, int((end - start).total_seconds())),
    )


stringPattern = r'"(.*?)"'

if __name__ == '__main__':
    # get reading file paths
    readingFilePaths = [readingFile for readingFile in os.listdir('.') if readingFile.endswith('.rdat')]
    # open and set up workbook to save readings in
    outputWorkbook = Workbook()
    outputWorkbook.worksheets[0].title = 'readings'
    outputWorkbook.active = 0

    # for every readings data file
    for filePath in readingFilePaths:
        startDateTime = datetime(year=2019, month=3, day=12, hour=12, minute=30)
        journeyTime = 0
        readings = []

        # open the file and get the readings and journey time
        with open(filePath) as readingFile:
            for line in readingFile:
                if 'TIME:' in line:
                    journeyTime = datetime.strptime(re.findall(stringPattern, line)[0], '%H:%M')

                if ('trkpt' in line) and ('lat' in line) and ('lon' in line):
                    readings.append(re.findall(stringPattern, line))

        # write out the readings with time stamps
        row = 1
        latColumn = first_empty_column(outputWorkbook.active)
        lonColumn = next_column_letter(latColumn)
        timeStampColumn = next_column_letter(lonColumn)
        outputWorkbook.active[latColumn + str(row)].value = "Lat"
        outputWorkbook.active[lonColumn + str(row)].value = "Lon"
        outputWorkbook.active[timeStampColumn + str(row)].value = "stamp"
        timestampDelta = timedelta(
            hours=journeyTime.hour / len(readings),
            minutes=journeyTime.minute / len(readings),
        )
        lastTimestamp = startDateTime
        for reading in readings:
            row += 1
            lastTimestamp += timestampDelta
            outputWorkbook.active[latColumn + str(row)].value = reading[0]
            outputWorkbook.active[lonColumn + str(row)].value = reading[1]
            outputWorkbook.active[timeStampColumn + str(row)].value = lastTimestamp.strftime("%s")

    outputWorkbook.save('output.xlsx')
    print('done')
